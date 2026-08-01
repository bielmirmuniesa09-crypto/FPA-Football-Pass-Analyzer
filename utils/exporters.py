from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import networkx as nx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.match_data import MatchData
from utils.graph_utils import (
    calcular_centralitats,
    calcular_mides_nodes,
    crear_graf_passades,
    crear_matriu_passades,
    obtenir_estadistiques_completes,
    obtenir_etiquetes_nodes,
    obtenir_pesos_arestes,
    obtenir_posicions_graf,
)


# =========================================================
# CARPETES I NOMS D'ARXIU
# =========================================================

def obtenir_carpeta_exports() -> Path:
    """
    Retorna la carpeta d'exportacions i la crea si no existeix.
    """

    carpeta_projecte = Path(__file__).resolve().parent.parent
    carpeta_exports = carpeta_projecte / "exports"

    carpeta_exports.mkdir(
        parents=True,
        exist_ok=True,
    )

    return carpeta_exports


def netejar_text_nom_arxiu(text: str) -> str:
    """
    Converteix un text en un nom d'arxiu segur.
    """

    text_net = text.strip().lower()

    substitucions = {
        "à": "a",
        "á": "a",
        "è": "e",
        "é": "e",
        "í": "i",
        "ï": "i",
        "ò": "o",
        "ó": "o",
        "ú": "u",
        "ü": "u",
        "ç": "c",
        "·": "_",
        " ": "_",
        "/": "-",
        "\\": "-",
        ":": "-",
        ";": "-",
        ",": "",
        ".": "",
        "'": "",
        '"': "",
    }

    for original, substitut in substitucions.items():
        text_net = text_net.replace(
            original,
            substitut,
        )

    caracters_permesos = []

    for caracter in text_net:
        if caracter.isalnum() or caracter in {
            "_",
            "-",
        }:
            caracters_permesos.append(
                caracter
            )

    resultat = "".join(
        caracters_permesos
    )

    return resultat or "partit"


def crear_nom_base(
    dades_partit: MatchData,
) -> str:
    """
    Crea un nom base únic per als arxius exportats.
    """

    equip = netejar_text_nom_arxiu(
        dades_partit.equip
    )

    rival = netejar_text_nom_arxiu(
        dades_partit.rival
    )

    marca_temps = datetime.now().strftime(
        "%Y%m%d_%H%M%S"
    )

    return (
        f"{equip}_vs_{rival}_"
        f"{marca_temps}"
    )


# =========================================================
# ESTILS D'EXCEL
# =========================================================

COLOR_CAPÇALERA = "0D2A22"
COLOR_VERD = "22C55E"
COLOR_VERMELL = "EF4444"
COLOR_GROC = "FACC15"
COLOR_BLAU_CLAR = "D9EAF7"
COLOR_BLANC = "FFFFFF"
COLOR_GRIS = "E5E7EB"


def aplicar_estil_capçalera(
    cel·la,
) -> None:
    """
    Aplica estil a una cel·la de capçalera.
    """

    cel·la.fill = PatternFill(
        fill_type="solid",
        fgColor=COLOR_CAPÇALERA,
    )

    cel·la.font = Font(
        bold=True,
        color=COLOR_BLANC,
    )

    cel·la.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )


def aplicar_estil_titol(
    cel·la,
) -> None:
    """
    Aplica estil a un títol de full.
    """

    cel·la.fill = PatternFill(
        fill_type="solid",
        fgColor=COLOR_VERD,
    )

    cel·la.font = Font(
        bold=True,
        size=16,
        color="04130E",
    )

    cel·la.alignment = Alignment(
        horizontal="center",
        vertical="center",
    )


def ajustar_amplades_columnes(
    full,
    amplada_maxima: int = 35,
) -> None:
    """
    Ajusta automàticament l'amplada de les columnes.
    """

    for columna in full.columns:
        longitud_maxima = 0

        lletra_columna = get_column_letter(
            columna[0].column
        )

        for cel·la in columna:
            valor = cel·la.value

            if valor is None:
                continue

            longitud = len(
                str(valor)
            )

            if longitud > longitud_maxima:
                longitud_maxima = longitud

        amplada = min(
            longitud_maxima + 3,
            amplada_maxima,
        )

        full.column_dimensions[
            lletra_columna
        ].width = max(
            amplada,
            10,
        )


# =========================================================
# EXPORTACIÓ EXCEL
# =========================================================

def exportar_excel(
    dades_partit: MatchData,
    ruta_sortida: str | Path | None = None,
) -> Path:
    """
    Exporta totes les dades del partit a un llibre Excel.

    Fulls inclosos:
    - Resum;
    - Jugadors;
    - Historial;
    - Matriu;
    - Centralitats.
    """

    carpeta_exports = obtenir_carpeta_exports()

    if ruta_sortida is None:
        nom_base = crear_nom_base(
            dades_partit
        )

        ruta = carpeta_exports / (
            f"{nom_base}.xlsx"
        )

    else:
        ruta = Path(
            ruta_sortida
        )

    llibre = Workbook()

    full_resum = llibre.active
    full_resum.title = "Resum"

    crear_full_resum(
        full_resum,
        dades_partit,
    )

    full_jugadors = llibre.create_sheet(
        "Jugadors"
    )

    crear_full_jugadors(
        full_jugadors,
        dades_partit,
    )

    full_historial = llibre.create_sheet(
        "Historial"
    )

    crear_full_historial(
        full_historial,
        dades_partit,
    )

    full_matriu = llibre.create_sheet(
        "Matriu"
    )

    crear_full_matriu(
        full_matriu,
        dades_partit,
    )

    full_centralitats = llibre.create_sheet(
        "Centralitats"
    )

    crear_full_centralitats(
        full_centralitats,
        dades_partit,
    )

    llibre.save(
        ruta
    )

    return ruta


def crear_full_resum(
    full,
    dades_partit: MatchData,
) -> None:
    """
    Crea el full de resum general.
    """

    full.merge_cells(
        "A1:D1"
    )

    full["A1"] = (
        "FOOTBALL PASS ANALYZER"
    )

    aplicar_estil_titol(
        full["A1"]
    )

    resum = dades_partit.resum_partit()

    files = [
        (
            "Equip",
            resum["equip"],
        ),
        (
            "Rival",
            resum["rival"],
        ),
        (
            "Competició",
            resum["competicio"],
        ),
        (
            "Data",
            resum["data_partit"],
        ),
        (
            "Sistema",
            resum["sistema"],
        ),
        (
            "Condició",
            resum["local_visitant"],
        ),
        (
            "Passades totals",
            resum["total"],
        ),
        (
            "Passades bones",
            resum["bones"],
        ),
        (
            "Passades dolentes",
            resum["dolentes"],
        ),
        (
            "Precisió global",
            resum["precisio"] / 100,
        ),
    ]

    full["A3"] = "CONCEPTE"
    full["B3"] = "VALOR"

    aplicar_estil_capçalera(
        full["A3"]
    )
    aplicar_estil_capçalera(
        full["B3"]
    )

    for fila, (
        concepte,
        valor,
    ) in enumerate(
        files,
        start=4,
    ):
        full.cell(
            row=fila,
            column=1,
            value=concepte,
        )

        full.cell(
            row=fila,
            column=2,
            value=valor,
        )

    full["B13"].number_format = "0.00%"

    full.freeze_panes = "A4"

    ajustar_amplades_columnes(
        full
    )


def crear_full_jugadors(
    full,
    dades_partit: MatchData,
) -> None:
    """
    Crea el full d'estadístiques individuals.
    """

    capçaleres = [
        "Dorsal",
        "Jugador",
        "Posició",
        "Intentades",
        "Bones",
        "Dolentes",
        "Rebudes",
        "Precisió",
    ]

    for columna, text in enumerate(
        capçaleres,
        start=1,
    ):
        cel·la = full.cell(
            row=1,
            column=columna,
            value=text,
        )

        aplicar_estil_capçalera(
            cel·la
        )

    estadistiques = (
        dades_partit
        .estadistiques_tots_jugadors()
    )

    for fila, jugador in enumerate(
        estadistiques,
        start=2,
    ):
        valors = [
            jugador["dorsal"],
            jugador["jugador"],
            jugador["posicio"],
            jugador["intentades"],
            jugador["bones"],
            jugador["dolentes"],
            jugador["rebudes"],
            jugador["precisio"] / 100,
        ]

        for columna, valor in enumerate(
            valors,
            start=1,
        ):
            full.cell(
                row=fila,
                column=columna,
                value=valor,
            )

        full.cell(
            row=fila,
            column=8,
        ).number_format = "0.00%"

    full.freeze_panes = "A2"

    ajustar_amplades_columnes(
        full
    )


def crear_full_historial(
    full,
    dades_partit: MatchData,
) -> None:
    """
    Crea el full amb totes les passades en ordre cronològic.
    """

    capçaleres = [
        "Número",
        "Origen",
        "Destí",
        "Resultat",
    ]

    for columna, text in enumerate(
        capçaleres,
        start=1,
    ):
        cel·la = full.cell(
            row=1,
            column=columna,
            value=text,
        )

        aplicar_estil_capçalera(
            cel·la
        )

    historial = dades_partit.historial_passades()

    for fila, passada in enumerate(
        historial,
        start=2,
    ):
        full.cell(
            row=fila,
            column=1,
            value=passada["numero"],
        )

        full.cell(
            row=fila,
            column=2,
            value=passada["origen"],
        )

        full.cell(
            row=fila,
            column=3,
            value=passada["desti"],
        )

        cel·la_resultat = full.cell(
            row=fila,
            column=4,
            value=passada["resultat"].upper(),
        )

        color = (
            COLOR_VERD
            if passada["resultat"] == "bona"
            else COLOR_VERMELL
        )

        cel·la_resultat.fill = PatternFill(
            fill_type="solid",
            fgColor=color,
        )

        cel·la_resultat.font = Font(
            bold=True,
            color=COLOR_BLANC,
        )

    full.freeze_panes = "A2"

    ajustar_amplades_columnes(
        full
    )


def crear_full_matriu(
    full,
    dades_partit: MatchData,
) -> None:
    """
    Crea el full amb la matriu de passades bones.
    """

    matriu = crear_matriu_passades(
        dades_partit
    )

    noms_curts = []

    for index, nom in enumerate(
        dades_partit.jugadors
    ):
        dorsal = (
            dades_partit.dorsals[index]
            if index < len(dades_partit.dorsals)
            else str(index + 1)
        )

        noms_curts.append(
            f"{dorsal}. {nom}"
        )

    full.cell(
        row=1,
        column=1,
        value="Origen / Destí",
    )

    aplicar_estil_capçalera(
        full.cell(
            row=1,
            column=1,
        )
    )

    for index, nom in enumerate(
        noms_curts,
        start=2,
    ):
        cel·la = full.cell(
            row=1,
            column=index,
            value=nom,
        )

        aplicar_estil_capçalera(
            cel·la
        )

    for fila_index, nom in enumerate(
        noms_curts,
        start=2,
    ):
        cel·la_nom = full.cell(
            row=fila_index,
            column=1,
            value=nom,
        )

        aplicar_estil_capçalera(
            cel·la_nom
        )

        fila_matriu = matriu[
            fila_index - 2
        ]

        for columna_index, valor in enumerate(
            fila_matriu,
            start=2,
        ):
            cel·la = full.cell(
                row=fila_index,
                column=columna_index,
                value=valor,
            )

            cel·la.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

    full.freeze_panes = "B2"

    ajustar_amplades_columnes(
        full,
        amplada_maxima=22,
    )


def crear_full_centralitats(
    full,
    dades_partit: MatchData,
) -> None:
    """
    Crea el full amb les mesures de teoria de grafs.
    """

    capçaleres = [
        "Dorsal",
        "Jugador",
        "Degree",
        "In-Degree",
        "Out-Degree",
        "Betweenness",
        "Closeness",
        "PageRank",
    ]

    for columna, text in enumerate(
        capçaleres,
        start=1,
    ):
        cel·la = full.cell(
            row=1,
            column=columna,
            value=text,
        )

        aplicar_estil_capçalera(
            cel·la
        )

    estadistiques = obtenir_estadistiques_completes(
        dades_partit
    )

    for fila, jugador in enumerate(
        estadistiques,
        start=2,
    ):
        valors = [
            jugador["dorsal"],
            jugador["jugador"],
            jugador["degree"],
            jugador["in_degree"],
            jugador["out_degree"],
            jugador["betweenness"],
            jugador["closeness"],
            jugador["pagerank"],
        ]

        for columna, valor in enumerate(
            valors,
            start=1,
        ):
            cel·la = full.cell(
                row=fila,
                column=columna,
                value=valor,
            )

            if columna >= 3:
                cel·la.number_format = "0.0000"

    full.freeze_panes = "A2"

    ajustar_amplades_columnes(
        full
    )


# =========================================================
# EXPORTACIÓ DEL GRAF EN PNG
# =========================================================

def exportar_graf_png(
    dades_partit: MatchData,
    ruta_sortida: str | Path | None = None,
) -> Path:
    """
    Genera una imatge PNG del graf de passades.
    """

    carpeta_exports = obtenir_carpeta_exports()

    if ruta_sortida is None:
        nom_base = crear_nom_base(
            dades_partit
        )

        ruta = carpeta_exports / (
            f"{nom_base}_graf.png"
        )

    else:
        ruta = Path(
            ruta_sortida
        )

    graf = crear_graf_passades(
        dades_partit
    )

    posicions = obtenir_posicions_graf(
        dades_partit
    )

    etiquetes = obtenir_etiquetes_nodes(
        dades_partit
    )

    pesos = obtenir_pesos_arestes(
        dades_partit
    )

    mides_dict = calcular_mides_nodes(
        dades_partit
    )

    mides_nodes = [
        mides_dict.get(
            node,
            1200,
        )
        for node in graf.nodes
    ]

    amplades_arestes = [
        max(
            1.0,
            pesos.get(
                (
                    origen,
                    desti,
                ),
                1,
            ) * 0.7,
        )
        for origen, desti in graf.edges
    ]

    figura = plt.figure(
        figsize=(14, 10)
    )

    ax = figura.add_subplot(
        111
    )

    ax.set_title(
        (
            f"Graf de passades — "
            f"{dades_partit.equip} vs "
            f"{dades_partit.rival}"
        ),
        fontsize=18,
        fontweight="bold",
        pad=20,
    )

    nx.draw_networkx_nodes(
        graf,
        posicions,
        node_size=mides_nodes,
        ax=ax,
    )

    nx.draw_networkx_edges(
        graf,
        posicions,
        width=amplades_arestes,
        arrows=True,
        arrowsize=20,
        connectionstyle="arc3,rad=0.08",
        ax=ax,
    )

    nx.draw_networkx_labels(
        graf,
        posicions,
        labels=etiquetes,
        font_size=9,
        font_weight="bold",
        ax=ax,
    )

    etiquetes_arestes = {
        (
            origen,
            desti,
        ): pes
        for (
            origen,
            desti,
        ), pes in pesos.items()
    }

    nx.draw_networkx_edge_labels(
        graf,
        posicions,
        edge_labels=etiquetes_arestes,
        font_size=8,
        rotate=False,
        ax=ax,
    )

    ax.axis(
        "off"
    )

    figura.tight_layout()

    figura.savefig(
        ruta,
        dpi=220,
        bbox_inches="tight",
    )

    plt.close(
        figura
    )

    return ruta


# =========================================================
# EXPORTACIÓ CSV
# =========================================================

def exportar_historial_csv(
    dades_partit: MatchData,
    ruta_sortida: str | Path | None = None,
) -> Path:
    """
    Exporta l'historial de passades a CSV.
    """

    import csv

    carpeta_exports = obtenir_carpeta_exports()

    if ruta_sortida is None:
        nom_base = crear_nom_base(
            dades_partit
        )

        ruta = carpeta_exports / (
            f"{nom_base}_historial.csv"
        )

    else:
        ruta = Path(
            ruta_sortida
        )

    historial = dades_partit.historial_passades()

    with ruta.open(
        mode="w",
        encoding="utf-8-sig",
        newline="",
    ) as arxiu:
        escriptor = csv.writer(
            arxiu,
            delimiter=";",
        )

        escriptor.writerow(
            [
                "Número",
                "Origen",
                "Destí",
                "Resultat",
            ]
        )

        for passada in historial:
            escriptor.writerow(
                [
                    passada["numero"],
                    passada["origen"],
                    passada["desti"],
                    passada["resultat"],
                ]
            )

    return ruta


# =========================================================
# EXPORTACIÓ CONJUNTA
# =========================================================

def exportar_tots_els_formats(
    dades_partit: MatchData,
) -> dict[str, Path]:
    """
    Exporta el partit a Excel, PNG i CSV.
    """

    ruta_excel = exportar_excel(
        dades_partit
    )

    ruta_png = exportar_graf_png(
        dades_partit
    )

    ruta_csv = exportar_historial_csv(
        dades_partit
    )

    return {
        "excel": ruta_excel,
        "png": ruta_png,
        "csv": ruta_csv,
    }